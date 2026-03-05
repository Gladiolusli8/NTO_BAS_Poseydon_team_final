from openpyxl import load_workbook
EXCEL_FILE = "drone_data.xlsx"

def add_data_to_excel(filename, measure_id, roll, pitch, yaw, val1=0, val2=0):
    """
    Добавляет строку в Excel файл, начиная с 3-й строки.
    Структура: ID, Крен, Тангаж, Рысканье, Доп1, Доп2
    """
    try:
        # Загружаем книгу
        wb = load_workbook(filename)
        ws = wb.active

        # Ищем первую пустую строку, начиная с 3-й
        row = 3
        while ws.cell(row=row, column=1).value is not None:
            row += 1

        # Записываем данные
        ws.cell(row=row, column=1).value = measure_id  # Номер измерения
        ws.cell(row=row, column=2).value = round(roll, 2)  # Крен
        ws.cell(row=row, column=3).value = round(pitch, 2)  # Тангаж
        ws.cell(row=row, column=4).value = round(yaw, 2)  # Рысканье
        ws.cell(row=row, column=5).value = val1  # Доп столбец 1
        ws.cell(row=row, column=6).value = val2  # Доп столбец 2

        # Сохраняем файл
        wb.save(filename)
        print(f"Данные записаны в строку {row}")

    except PermissionError:
        print("ОШИБКА: Файл Excel открыт в другой программе! Закройте его перед запуском.")
    except FileNotFoundError:
        print(f"ОШИБКА: Файл {filename} не найден. Создайте шаблон.")
    except Exception as e:
        print(f"Произошла ошибка при записи: {e}")
add_data_to_excel(EXCEL_FILE, 1, 104, 3349, 81, 23, 93)